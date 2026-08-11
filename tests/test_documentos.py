"""Testes do Gabinete Designer: parser Markdown + geração Word/PDF/Excel."""

import io

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.dependencies import get_current_user
from app.schemas.auth import AuthUser
from app.services.documentos import (
    Timbre,
    gerar_docx,
    gerar_pdf,
    gerar_xlsx,
    parse_markdown,
)

_MD = """# Notificação Extrajudicial

Prezado **condômino**, vimos por meio desta notificar sobre o débito.

## Fundamentos
- Art. 1.336 do Código Civil
- Convenção do condomínio

| Mês | Valor |
| --- | --- |
| Janeiro | 500 |
| Fevereiro | 500 |
"""


def test_parse_markdown_reconhece_blocos() -> None:
    blocos = parse_markdown(_MD)
    tipos = [t for t, _ in blocos]
    assert "h1" in tipos and "h2" in tipos
    assert "lista" in tipos and "paragrafo" in tipos
    tabela = next(c for t, c in blocos if t == "tabela")
    # cabeçalho + 2 linhas (separador '---' descartado)
    assert tabela[0] == ["Mês", "Valor"]
    assert ["Janeiro", "500"] in tabela
    assert len(tabela) == 3


def test_gerar_docx_valido() -> None:
    dados = gerar_docx("Notificação", _MD, escritorio="Vale Advocacia")
    assert dados[:2] == b"PK"  # zip/docx
    assert len(dados) > 1000


def test_gerar_pdf_valido() -> None:
    dados = gerar_pdf("Notificação", _MD, escritorio="Vale Advocacia")
    assert dados[:4] == b"%PDF"
    assert len(dados) > 500


def test_gerar_pdf_lida_com_unicode_exotico() -> None:
    # em-dash, aspas curvas e emoji não podem quebrar a geração (fonte latin-1)
    dados = gerar_pdf("Teste", "Bloco — “aspas” e emoji 🎯 conforme art. 5º.")
    assert dados[:4] == b"%PDF"


def test_timbre_completo_gera_docx_e_pdf() -> None:
    timbre = Timbre(
        nome="Jales Advogados",
        subtitulo="OAB/GO 12.345 · Direito Condominial",
        cor="#2E4739",
        rodape="Documento confidencial — uso interno.",
    )
    docx = gerar_docx("Notificação", _MD, timbre=timbre)
    pdf = gerar_pdf("Notificação", _MD, timbre=timbre)
    assert docx[:2] == b"PK" and len(docx) > 1000
    assert pdf[:4] == b"%PDF" and len(pdf) > 500


def test_timbre_cor_invalida_nao_quebra() -> None:
    # cor malformada cai no padrão, sem estourar
    pdf = gerar_pdf("T", _MD, timbre=Timbre(nome="X", cor="banana"))
    assert pdf[:4] == b"%PDF"


def test_gerar_xlsx_usa_a_tabela() -> None:
    dados = gerar_xlsx("Inadimplência", _MD)
    assert dados[:2] == b"PK"
    wb = load_workbook(io.BytesIO(dados))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["Mês", "Valor"]
    assert ws["A2"].value == "Janeiro"


def test_exportar_endpoint_docx(client: TestClient) -> None:
    client.app.dependency_overrides[get_current_user] = lambda: AuthUser(id="u1", email="x@y")
    try:
        r = client.post(
            "/api/documentos/exportar",
            json={
                "conteudo": _MD,
                "formato": "docx",
                "titulo": "Notificação",
                "escritorio": "Vale",
            },
        )
        assert r.status_code == 200
        assert "wordprocessingml" in r.headers["content-type"]
        assert r.content[:2] == b"PK"
        assert "attachment" in r.headers["content-disposition"]
    finally:
        client.app.dependency_overrides.clear()


def test_exportar_endpoint_exige_login(client: TestClient) -> None:
    r = client.post(
        "/api/documentos/exportar",
        json={"conteudo": "x", "formato": "pdf", "titulo": "T"},
    )
    assert r.status_code in (401, 403)
